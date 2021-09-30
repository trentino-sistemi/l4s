# This file is part of Lod4Stat.
#
# Copyright (C) 2014 Provincia autonoma di Trento
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""
Actions for l4s project.
"""

from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.templatetags.static import static
from django.core.exceptions import ObjectDoesNotExist
from explorer.models import Query
from l4s import settings
from .rdf import rdf_report
from .sdmx import sdmx_report
from tempfile import NamedTemporaryFile
from pandas import ExcelWriter
from web.pyjstat import to_json_stat
from web.reconcilation import reconciles_data_frame
from web.utils import unpivot,\
    has_data_frame_multi_level_columns, \
    get_color
from web.statistical_secret import load_data_frame
from xlrd import open_workbook
from xlwt import Workbook as XWorkbook
from xlwt import easyxf, add_palette_colour
from openpyxl.styles import NamedStyle, Color, PatternFill, Font
from openpyxl.styles import Alignment as OAlignment
from openpyxl import Workbook as OWorkbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from web import arial10
import ast
import six
import calendar
from io import StringIO
import shutil

max_length_filename = 150 # sotto linux dipende dal file system ... sotto ntfs massimo 240 compreso path ... ma il path non lo so .....per ora metto 150 prudenziale

def new_xlwt_colored_workbook():
    """
    Get an Xlwt Workbook with the custom colors.

    :return: xlwt Woorkbook
    """
    new_workbook = XWorkbook(encoding="UTF-8")
    new_workbook.set_colour_RGB(0x21, 139, 31, 63)
    new_workbook.set_colour_RGB(0x22, 255, 255, 255)
    new_workbook.set_colour_RGB(0x23, 31, 85, 111)
    add_palette_colour("custom_colour", 0x21)
    add_palette_colour("white", 0x22)
    add_palette_colour("blue", 0x23)
    return new_workbook


def generate_report_action_csv(request):
    """
    Generate CSV file from SQL query.
    :param request: Django request.
    """
    def generate_report(title):
        """
        Generate Csv file from query.

        :param title: The query title.
        :return: Response with CSV attachment.
        """
        df = load_data_frame(request)
        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        extension = '.csv'
        separator = ';'
        filename = '%s%s' % (title, extension)
        content_type = 'text/csv'
        out_stream = StringIO()
        df.to_csv(out_stream, sep=separator, index=True)
        # Setup response
        response = HttpResponse(out_stream.getvalue())
        response["Content-Type"] = content_type
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    return generate_report


def generate_report_action_xls(request):
    """
    Generate Excel 1997 file from SQL query.

    :param request: Django request.
    :return: Response with Excel 1997 attachment.
    """

    def add_header_and_footer(file_name, title, description, show_legend, table_description):
        """
        Add header and footer to excel file.

        :param file_name: Excel file name.
        :param title: Query tile.
        :param description: Query description.
        """
        workbook = open_workbook(file_name)
        sheet = workbook.sheet_by_index(0)
        header_len = 5
        sheet_rows = sheet.nrows

        new_workbook = new_xlwt_colored_workbook()

        head_cfg = 'pattern: pattern solid, fore_colour custom_colour;'
        head_cfg += 'font: colour white, bold True;'
        head_cfg += 'alignment: horizontal left, vertical top, wrap true;'
        head_cell = easyxf(head_cfg)

        body_cfg = 'font: colour blue;'
        body_cfg += 'alignment: vertical top, wrap true;'
        body_cell = easyxf(body_cfg)

        ast_cfg = 'font: colour blue;'
        ast_cfg += 'alignment: vertical top, horizontal right, wrap true;'
        ast_cell = easyxf(ast_cfg)

        new_sheet = new_workbook.add_sheet("Lod4Stat", cell_overwrite_ok=True)

        table_description_label = str(_("Argomento"))
        new_sheet.write(0, 0, table_description_label, head_cell)
        new_sheet.write(0, 1, table_description, head_cell)
        new_sheet.write_merge(0, 0, 1, header_len, table_description, head_cell)

        title_label = str(_("Title"))
        new_sheet.write(1, 0, title_label, head_cell)
        new_sheet.write(1, 1, title, head_cell)
        new_sheet.write_merge(1, 1, 1, header_len, title, head_cell)


        line_num = 2
        if description is not None:
            description_label = str(_("Description"))
            s = StringIO(description)
            char_per_cell = 10
            for line in s:
                line_num += 1
                if len(line) > char_per_cell * header_len:
                    add = (len(line) // (char_per_cell * header_len)) + 1
                    line_num += add

            new_sheet.write(2, 0, description_label, head_cell)
            new_sheet.write_merge(2, line_num, 0, 0, description_label, head_cell)
            new_sheet.write(2, 1, description, head_cell)
            new_sheet.write_merge(2, line_num, 1, header_len, description, head_cell)

        #print "show_legend ", show_legend

        if show_legend == 'True':
            legend = "%s (%s)" % (settings.LEGEND, settings.DL_ART)
            legend_cells = 4
            new_sheet.write(line_num + 2, 0, legend, body_cell)
            new_sheet.write_merge(line_num + 2,
                                  line_num + 2,
                                  0,
                                  legend_cells,
                                  legend,
                                  body_cell)
            line_num += 2

        k = 2 + line_num
        max_widths = dict()
        default_width = 10
        for col in range(sheet.ncols):
            max_widths[col] = default_width
        #Copy rows from existing sheets
        for rows in range(0, sheet_rows):
            data = [sheet.cell_value(rows, col) for col in range(sheet.ncols)]
            for index, value in enumerate(data):
                #print index, value
                column_len = arial10.fit_width(value, False)
                #print column_len
                #print isinstance(value, unicode)
                if isinstance(value, str):
                    if value.isdigit():
                        #print "a"
                        #print "prima '",value,"'"
                        #print type(value)

                        if len(value) > 1 and value.startswith("0"): #sembra un baco della eval quando i numeri cominciano con 0
                            value = str(value)
                        else:
                            value = ast.literal_eval(value)

                        #print "dopo ", value
                        #print type(value)
                        new_sheet.write(rows+k, index, value, body_cell)
                    else:
                        #print "b"
                        #print value
                        value = value.strip()
                        if value.startswith("*"):
                            new_sheet.write(rows+k, index, value, ast_cell)
                        else:
                            new_sheet.write(rows+k, index, value, body_cell)
                else:
                    new_sheet.write(rows+k, index, value, body_cell)
                if column_len > max_widths[index]:
                    max_widths[index] = column_len
        # Adjust column width.
        for col in max_widths:
            new_sheet.col(col).width = round(max_widths[col]).__int__() + 1
            new_sheet.col(col)

        k = k + sheet_rows + 1

        stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'

        """
        if settings.DEBUG:
            stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'
        else:
            stat_bitmap = static('/img/testata_Statistica.bmp')
        """

        new_sheet.insert_bitmap(stat_bitmap, k, 0)
        new_workbook.save(file_name)

    def generate_report(title, description):
        """Generate Excel  1997 file from query.

        :param title: Query title.
        :param description: Query description.
        :return: Response with Excel 1997 attachment.
        """
        df = load_data_frame(request)

        # Limit the columns to the maximum allowed in Excel 97.
        max_length = 255
        index_len = len(df.index.names)

        lim_df = df.drop(df.columns[max_length - index_len - 1:len(df.columns) - 1], axis=1)

        extension = '.xls'
        engine = 'xlwt'
        encoding = 'utf-8'
        content_type = 'application/vnd.ms-excel'
        # Add content and return response
        f = NamedTemporaryFile(suffix=extension)
        ew = ExcelWriter(f.name, engine=engine, encoding=encoding)

        #print lim_df.to_string()
        #print f.name

        lim_df.to_excel(ew)
        ew.save()


        #shutil.copyfile(f.name, 'manuel.xls')

        show_legend = request.POST.get('show_legend', '') or request.GET.get(
            'show_legend', '')
        table_description = request.POST.get('table_description',
                                             '') or request.GET.get(
                                                 'table_description', '')

        add_header_and_footer(f.name, title, description, show_legend, table_description)

        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        filename = '%s%s' % (title, extension)

        # Setup response
        data = f.read()

        response = HttpResponse(data)
        response["Content-Type"] = content_type
        response["Content-status_code"] = 200
        response['Content-Transfer-Encoding'] = 'binary'
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    return generate_report


def generate_report_action_xlsx(request):
    """
    Generate Excel 2007 file from SQL query.

    :param request: Django request.
    :return: Response with Excel 2007 attachment.
    """

    def add_header_and_footer(file_name, title, description, show_legend, table_description):
        """
        Add header and footer to excel file.

        :param file_name: Excel file name.
        :param title: Query tile.
        :param description: Query description.
        """
        workbook = load_workbook(filename=file_name, read_only=True)
        sheet = workbook.active
        new_workbook = OWorkbook()
        header_len = 5
        new_sheet = new_workbook.active

        title_label = str(_("Title"))
        table_description_label = str(_("Argomento"))

        line_num = 2

        r_alignment = OAlignment(horizontal='right')
        l_alignment = OAlignment(horizontal='left')
        top_and_left_alignment = OAlignment(vertical='top', horizontal='left', wrapText=True)

        header_fill = PatternFill(patternType='solid',
                                  fgColor=Color('8B1F3F'))

        header_font = Font(color="FFFFFF", bold=True)
        header_style = NamedStyle(name='header', fill=header_fill, font=header_font, alignment=top_and_left_alignment)

        body_font = Font(color="1F556F")

        r_style = NamedStyle(name='r', font=body_font, alignment=r_alignment)
        body_style = NamedStyle(name='body', font=body_font)

        cell = new_sheet.cell(row=1, column=1)
        cell.value = table_description_label
        cell.style = header_style

        cell = new_sheet.cell(row=1, column=2)
        cell.value = table_description
        cell.style = header_style

        new_sheet.merge_cells(start_row=1, start_column=2,
                              end_row=1, end_column=header_len+1)


        cell = new_sheet.cell(row=2, column=1)
        cell.value = title_label
        cell.style = header_style

        cell = new_sheet.cell(row=2, column=2)
        cell.value = title
        cell.style = header_style

        new_sheet.merge_cells(start_row=2, start_column=2,
                              end_row=2, end_column=header_len+1)

        if description is not None:
            description_label = str(_("Description"))
            s = StringIO(description)
            char_per_cell = 10
            for line in s:
                line_num += 1
                if len(line) > char_per_cell * header_len:
                    add = (len(line) // (char_per_cell * header_len)) + 1
                    line_num += add

            cell = new_sheet.cell(row=3, column=1)
            cell.style = header_style
            cell.value = description_label
            new_sheet.merge_cells(start_row=3,
                                  start_column=1,
                                  end_row=line_num + 1,
                                  end_column=1)

            cell = new_sheet.cell(row=3, column=2)
            cell.style = header_style
            cell.value = description

            new_sheet.merge_cells(start_row=3,
                                  start_column=2,
                                  end_row=line_num + 1,
                                  end_column=header_len + 1)

        if show_legend == 'True':
            line_num += 3
            cell = new_sheet.cell(row=line_num, column=1)
            cell.value = "%s (%s)" % (settings.LEGEND, settings.DL_ART)
            cell.style = body_style
            legend_cells = 4
            new_sheet.merge_cells(start_row=line_num,
                                  start_column=1,
                                  end_row=line_num,
                                  end_column=legend_cells + 1)

        k = 1 + line_num
        max_widths = dict()

        default_width = 10
        for r, row in enumerate(sheet.iter_rows()):
            for c, cell in enumerate(row):
                max_widths[c] = default_width
            break

        # Copy rows from existing sheet.
        for r, row in enumerate(sheet.iter_rows()):
            for c, cell in enumerate(row):
                value = cell.value
                if value is None:
                    continue
                cell = new_sheet.cell(row=r + k + 1, column=c + 1)
                if isinstance(value, six.string_types):
                    value = value.strip()
                    if value.startswith("*") or value.isdigit():
                        cell.style = r_style
                    else:
                        cell.style = body_style
                else:
                    cell.style = body_style
                cell.value = value
                r_size = arial10.fit_width(str(value), False) // 255
                column_len = round(r_size)
                if r >= 1 and column_len > max_widths[c]:
                        max_widths[c] = column_len

        for i in max_widths:
            column_letter = get_column_letter(i+1)
            new_sheet.column_dimensions[column_letter].width = max_widths[i]

        sheet_rows = len(list(new_sheet.rows))
        k = k + sheet_rows - 1

        stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'

        """
        if settings.DEBUG:
            stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'
        else:
            stat_bitmap = static('/img/testata_Statistica.bmp')
        """

        stat_img = Image(stat_bitmap)
        stat_img.anchor = f'A{k}'

        #new_cell = new_sheet.cell(row=k, column=1)
        #stat_img.anchor(new_cell)
        new_sheet.add_image(stat_img)

        new_workbook.save(file_name)

    def generate_report(title, description):
        """Generate Excel 2007 file from query.

        :param title: Query title.
        :param description: Query description.
        :return: Response with Excel 2007 attachment.
        """
        extension = '.xlsx'
        engine = "openpyxl"
        encoding = 'utf-8'

        df = load_data_frame(request)

        # Add content and return response
        f = NamedTemporaryFile(suffix=extension)
        ew = ExcelWriter(f.name, engine=engine, options={'encoding': encoding})
        df.to_excel(ew)
        ew.save()

        show_legend = request.POST.get('show_legend', '') or request.GET.get(
            'show_legend', '')
        table_description = request.POST.get('table_description',
                                             '') or request.GET.get(
                                                 'table_description', '')

        add_header_and_footer(f.name, title, description, show_legend, table_description)

        # Setup response
        data = f.read()

        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        filename = '%s%s' % (title, extension)
        # Setup response
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(data, content_type=content_type)
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        # Add content and return response
        return response

    return generate_report


def generate_report_action_sdmx(request):
    """
     Generate SDMX file from SQL query.

    :param request: Django request.
    :return: Response with SDMX attachment.
    """

    def generate_report(title, sql):
        """
        Generate Sdmx file from query.

        :param title: The query title.
        :param sql: The query sql.
        :return: Response with Sdmx attachment.
        """
        df = load_data_frame(request)
        multi_index = has_data_frame_multi_level_columns(df)
        if multi_index:
            int_df = unpivot(df)
        else:
            int_df = df

        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        extension = '.sdmx'
        filename = '%s%s' % (title, extension)
        content_type = 'application/xml'
        res = sdmx_report(sql, int_df)
        # Setup response
        response = HttpResponse(res)
        response["content_type"] = content_type
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        return response

    return generate_report


def generate_report_action_json_stat(request):
    """
    Generate JSON-stat file from SQL query.

    :param request: Django request.
    :return: Response with CSV attachment.
    """
    def generate_report(title):
        """
        Generate JSON-stat file from query.

        :return: Response with JSON-stat attachment.
        """
        df = load_data_frame(request)
        sql = query_sql(request)
        #df = reconciles_data_frame(df, sql)

        multi_index = has_data_frame_multi_level_columns(df)
        if multi_index:
            int_df = unpivot(df)
            value = df.columns.levels[0][0]
        else:
            int_df = df
            value = df.columns[len(df.columns)-1]

        int_df = reconciles_data_frame(int_df, sql)
        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        extension = '.json'
        filename = '%s%s' % (title, extension)
        content_type = 'application/json'

        response = HttpResponse()
        response["content_type"] = content_type
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)

        if value is None:
            return response

        val = to_json_stat(int_df, value=value)
        # Setup response
        response.write(val)

        return response

    return generate_report


def generate_report_action_rdf(request):
    """
    Generate RDF file from SQL query.

    :param request: Django request.
    :return: Response with RDF attachment.
    """

    def generate_report(title, description, sql):
        """
        Generate Rdf file from query.

        :param title:  Query title.
        :param description: query description.
        :param sql: Query sql.
        :return: Response with RDF attachment.
        """
        df = load_data_frame(request)
        multi_index = has_data_frame_multi_level_columns(df)
        if multi_index:
            int_df = unpivot(df)
        else:
            int_df = df

        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        extension = '.xml'
        filename = '%s%s' % (title, extension)
        res = rdf_report(sql, title, description,
                         data_frame=int_df,
                         rdf_format='xml')
        # Setup response
        content_type = 'application/rdf+xml'
        response = HttpResponse(res)
        response["content_type"] = content_type

        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        return response

    return generate_report


def generate_report_action_turtle(request):
    """
    Generate TURTLE file from SQL query.

    :param request: Django request.
    :return: Response with TURTLE attachment.
    """

    def generate_report(title, description, sql):
        """
        Generate Turtle file from query.

        :param title:  Query title.
        :param description: query description.
        :param sql: Query sql.
        :return: Response with Turtle attachment.
        """
        df = load_data_frame(request)
        multi_index = has_data_frame_multi_level_columns(df)
        if multi_index:
            int_df = unpivot(df)
        else:
            int_df = df

        title = title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        extension = '.ttl'
        filename = '%s%s' % (title, extension)
        res = rdf_report(sql, title, description,
                         data_frame=int_df,
                         rdf_format='turtle')
        # Setup response
        content_type = 'text/turtle'
        response = HttpResponse(res)
        response["content_type"] = content_type
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        return response

    return generate_report


def query_title(request):
    """
    Get query title for the response.
    If it is specified in request then use it else try to take it from query
    title. If something go wrong use a fix string.

    :param request: Django request.
    :return: title
    """
    title = request.POST.get('title') or request.GET.get('title')
    if title is None:
        query_id = request.POST.get('id') or request.GET.get('id')
        try:
            query = Query.objects.get(id=query_id)
            title = query.title
        except ObjectDoesNotExist:
            title = 'result'
    return title


def query_description(request):
    """
    Get query title for the response.
    If it is specified in request then use it else try to take it from query
    title. If something go wrong use empty string.

    :param request:
    """
    description = request.POST.get('description') or request.GET.get(
        'description')
    if description is None:
        query_id = request.POST.get('id') or request.GET.get('id')
        try:
            query = Query.objects.get(id=query_id)
            description = query.description
        except ObjectDoesNotExist:
            description = ''
    return description


def query_sql(request):
    """
    Get sql query for the response.
    If it is specified in request then use it else try to take it from query
    title. If something go wrong use empty string.

    :param request:
    """
    sql = request.POST.get('sql') or request.GET.get('sql')
    if sql is None:
        query_id = request.POST.get('id') or request.GET.get('id')
        try:
            query = Query.objects.get(id=query_id)
            sql = query.sql
        except ObjectDoesNotExist:
            sql = ''
    return sql


def generate_usage_report_action_xls(request):

    def generate_report():
        extension = '.xls'
        f = NamedTemporaryFile(suffix=extension)

        content_type = 'application/vnd.ms-excel'
        # Setup response
        queries_s = request.POST.get('queries') or request.GET.get('queries')
        queries = eval("[%s]" % queries_s)[0]
        manual_request_s = request.POST.get(
            'manual_requests') or request.GET.get('manual_requests')
        manual_requests = eval("[%s]" % manual_request_s)[0]
        run_queries_auth_s = request.POST.get(
            'run_queries_auth') or request.GET.get('run_queries_auth')
        run_queries_auth = eval("[%s]" % run_queries_auth_s)[0]
        run_queries_anon_s = request.POST.get(
            'run_queries_anon') or request.GET.get('run_queries_anon')
        run_queries_anon = eval("[%s]" % run_queries_anon_s)[0]

        new_workbook = new_xlwt_colored_workbook()

        new_sheet = new_workbook.add_sheet("Lod4Stat", cell_overwrite_ok=True)
        head_cfg = 'pattern: pattern solid, fore_colour custom_colour;'
        head_cfg += 'font: colour white, bold True;'
        head_cfg += 'alignment: horizontal left, vertical top, wrap true;'
        head_cell = easyxf(head_cfg)
        header_len = 5
        us_title = str(_('Usage report'))
        year = request.POST.get('year') or request.GET.get('year')
        month = request.POST.get('month') or request.GET.get('month')
        us_title += " " + str(_('Year')) + " " + year
        if month != "None":
            month_name = calendar.month_name[ast.literal_eval(month)]
            us_title += ", " + str(_('Month')).lower() + " " + month_name

        body_cfg = 'font: colour blue;'
        body_cfg += 'alignment: vertical top, wrap true;'
        body_cell = easyxf(body_cfg)

        new_sheet.write(0, 0, us_title, head_cell)
        new_sheet.write_merge(0, 0, 0, header_len - 1, us_title, head_cell)

        max_widths = dict()
        default_width = 10
        for col in range(2):
            max_widths[col] = default_width

        start = 3
        title = str(_("Number of saved queries for user types"))
        new_sheet.write(start, 0, title, head_cell)
        new_sheet.write_merge(start, start, 0, 1, title, head_cell)
        start = start + 1

        title = str(_("User type"))
        new_sheet.write(start, 0, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[0]:
            max_widths[0] = column_len
        title = str(_("Saved queries"))
        new_sheet.write(start, 1, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[1]:
            max_widths[1] = column_len
        start = start + 1

        q = 0
        for q, query in enumerate(queries):
            new_sheet.write(q + start, 0, query[0], body_cell)
            column_len = arial10.fit_width(query[0], False)
            if column_len > max_widths[0]:
                max_widths[0] = column_len
            new_sheet.write(q + start, 1, query[1], body_cell)

        start = start + q + 3
        title = str(_("Number of manual requests for user types"))
        new_sheet.write(start, 0, title, head_cell)
        new_sheet.write_merge(start, start, 0, 1, title, head_cell)
        start = start + 1

        title = str(_("User type"))
        new_sheet.write(start, 0, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[0]:
            max_widths[0] = column_len
        title = str(_("Manual requests"))
        new_sheet.write(start, 1, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[1]:
            max_widths[1] = column_len

        start = start + 1

        for q, query in enumerate(manual_requests):
            new_sheet.write(q + start, 0, query[0], body_cell)
            new_sheet.write(q + start, 1, query[1], body_cell)
            column_len = arial10.fit_width(query[0], False)
            if column_len > max_widths[0]:
                max_widths[0] = column_len
            new_sheet.write(q + start, 1, query[1], body_cell)

        start = start + q + 3
        title = _("Number of executed queries for user types (logged users)")
        title = str(title)
        new_sheet.write(start, 0, title, head_cell)
        new_sheet.write_merge(start, start, 0, 1, title, head_cell)
        start = start + 1

        title = str(_("User type"))
        new_sheet.write(start, 0, title, head_cell)
        title = str(_("Run queries"))
        new_sheet.write(start, 1, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[1]:
            max_widths[1] = column_len

        start = start + 1

        for q, query in enumerate(run_queries_auth):
            new_sheet.write(q + start, 0, query[0], body_cell)
            new_sheet.write(q + start, 1, query[1], body_cell)
            column_len = arial10.fit_width(query[0], False)
            if column_len > max_widths[0]:
                max_widths[0] = column_len
            new_sheet.write(q + start, 1, query[1], body_cell)


        start = start + q + 3
        title = str(_("Number of executed queries (anonymous users)"))
        new_sheet.write(start, 0, title, head_cell)
        new_sheet.write_merge(start, start, 0, 1, title, head_cell)
        start = start + 1
        title = str(_("Run queries"))
        new_sheet.write(start, 0, title, head_cell)
        new_sheet.write_merge(start, start, 0, 1, title, head_cell)
        column_len = arial10.fit_width(title, False)
        if column_len > max_widths[0]:
            max_widths[0] = column_len

        start = start + 1

        for q, query in enumerate(run_queries_anon):
            new_sheet.write(q + start, 0, query[0], body_cell)
            new_sheet.write_merge(q + start,
                                  q + start,
                                  0,
                                  1,
                                  query[0],
                                  body_cell)
            column_len = arial10.fit_width(str(query[0]), False)
            if column_len > max_widths[0]:
                max_widths[0] = column_len

        # Adjust column width.
        for col in max_widths:
            new_sheet.col(col).width = round(max_widths[col]).__int__() + 1

        start = start + q + 3

        stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'

        """
        if settings.DEBUG:
            stat_bitmap = 'l4s/static/img/testata_Statistica.bmp'
        else:
            stat_bitmap = static('/img/testata_Statistica.bmp')
        """

        new_sheet.insert_bitmap(stat_bitmap, start, 0)

        new_workbook.save(f.name)

        title = us_title.strip().replace(" ", '_')

        if len(title) > max_length_filename:
            title = title[:max_length_filename]

        filename = '%s%s' % (title, extension)

        data = f.read()
        response = HttpResponse(data)
        response["Content-Type"] = content_type
        response['Content-Transfer-Encoding'] = 'binary'
        response[
            'Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    return generate_report
